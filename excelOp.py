#! /usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import Workbook


def writeAreaToExcel(area):
    wb = Workbook()
    ws = wb.active
    row = 1
    for k in area:
        ws.cell(row, 1, k)
        for i in range(9):
            ws.cell(row, i + 2, area[k][i])

        row += 1

    wb.save("area_first.xlsx")
